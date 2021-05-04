import openpyxl
import os
from selenium import webdriver


def start_webdriver():
    print('Please use the latest versions of the browsers.')
    print("Select the number of the browser( that you have installed in your PC) to test:")
    print("Chrome->1")
    print("Opera-2")
    print("Safari->3")
    print("Edge->4")
    print("Firefox->5")
    input1 = input()
    if input1 == '1':
        print('please download the lastest version of Chrome webdriver')
        driver_path = os.path.abspath("chromedriver.exe")
        driver = webdriver.Chrome(executable_path=driver_path)
    elif input1 == '2':
        print('please download the lastest version of Opera webdriver')
        driver_path = os.path.abspath("operadriver.exe")
        driver = webdriver.Opera(executable_path=driver_path)
    elif input1 == '3':
        print('check if is active the development version in Safari')
        driver = webdriver.Safari()
    elif input1 == '4':
        print('please download the lastest version of Edge webdriver')
        driver_path = os.path.abspath("msedgedriver.exe")
        driver = webdriver.Edge(driver_path)
    elif input1 == '5':
        print('please download the lastest version of Geckodriver')
        driver_path = os.path.abspath("geckodriver.exe")
        driver = webdriver.Firefox(executable_path=driver_path)
    return driver


def getrowcount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)


def getcolumncount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)


def readdata(file, sheetname, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnno).value


def writedata(file, sheetname, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)


def weeklist(text):  # mandar los que no van
    l = ['0', '1', '2', '3', '4', '5', '6']
    if text != "":
        a = text.split(",")
        print(a)
        for i in range(len(a)):
            c = a[i]
            if c[0] == ' ':
                a[i] = c[1:]
        # '0','1','2','3','4','5','6'
        print(a)
        for i in range(len(a)):
            l.remove(a[i])
            print(l)
            # a: arreglo q indica los valores que no estan en el arreglo original

        for j in range(len(l)):
            if l[j] == "0":
                l[j] = "Domingo"
            elif l[j] == "1":
                l[j] = "Lunes"
            elif l[j] == "2":
                l[j] = "Martes"
            elif l[j] == "3":
                l[j] = "Miércoles"
            elif l[j] == "4":
                l[j] = "Jueves"
            elif l[j] == "5":
                l[j] = "Viernes"
            else:
                l[j] = "Sábado"
        print(l)
        b = l
    else:
        b = []
    return b


def permissions(array):
    # AdminDefault, ClientViewer, Euler, ProviderViewer, ServicesEditor, TouroperatorDefault, UserViewer
    a = array.split(',')
    b = []
    for i in range(len(a)):
        c = a[i]
        if c[0] != ' ':
            b.append(c)
        else:
            b.append(c[1:])
    # print(b)
    return b
